from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import re
import keras
from keras import backend as K
from PIL import Image
from keras.models import load_model
import numpy as np
import h5py
import pandas as pd
from matplotlib import pyplot as plt
import os
import string
import io
from time import sleep
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import configparser

# CHRS = string.ascii_lowercase + string.digits  # 字符列表
CHRS = string.ascii_lowercase  # 字符列表
num_classes = 26  # 共要识别36个字符（所有小写字母+数字），即36类
batch_size = 128
epochs = 12

# 输入图片的尺寸
img_rows, img_cols = 12, 20
# 根据keras的后端是TensorFlow还是Theano转换输入形式
if K.image_data_format() == 'channels_first':
    input_shape = (1, img_rows, img_cols)
else:
    input_shape = (img_rows, img_cols, 1)

# 载入已经训练模型
model_path = os.path.split(os.path.realpath(__file__))[0] + '/model_test.h5'
model = load_model(model_path)
# model = load_model(r'E:\python\2017_9\model_test.h5')

def get_bin_table(threshold=98):
    """
    获取灰度转二值的映射table
    :param threshold: 边界值
    :return:
    """
    table = []
    for i in range(256):
        if i < threshold:
            table.append(0)
        else:
            table.append(1)

    return table

def handle_split_image(image):
    '''
    切割验证码，返回包含四个字符图像的列表
    '''
    # im = image.point(lambda i: i != 43, mode='1')
    # y_min, y_max = 0, 22 # im.height - 1 # 26
    # split_lines = [5,17,29,41,53]
    # ims = [im.crop([u, y_min, v, y_max]) for u, v in zip(split_lines[:-1], split_lines[1:])]
    # # w = w.crop(w.getbbox()) # 切掉白边 # 暂不需要

    # 转化到灰度图
    imgry = image.convert('L')
    # 二值化，采用阈值分割法，threshold为分割点
    out = imgry.point(get_bin_table(), '1')
    a = np.array(out)
    pd.DataFrame(a.sum(axis=0)).plot.line()
    # 画出每列的像素累计值
    # plt.imshow(a)  # 画出图像
    split_lines = [1, 13, 25, 30, 42, 48, 60]
    y_min, y_max = 0, 20
    image = [out.crop([u, y_min, v, y_max]) for u, v in zip(split_lines[:-1], split_lines[1:])]
    ims = []
    for i in range(0, 8, 2):
        plt.subplot(1, 4, i / 2 + 1)
        # print(i)
        if (i == 0):
            ims.append(image[int(i)])
            # plt.imshow(ims[int(i)], interpolation='none')
        else:
            ims.append(image[int(i - 1)])
            # plt.imshow(ims[int(i - 1)], interpolation='none')
    # print(ims)
    return ims

def predict(images):
    '''
    使用模型对四个字符的列表对应的验证码进行预测
    '''
    Y = []
    for i in range(4):
        im = images[i]
        test_input = np.concatenate(np.array(im))
        test_input = test_input.reshape(1, *input_shape)
        y_probs = model.predict(test_input)
        Y.append(CHRS[y_probs[0].argmax(-1)])
    return ''.join(Y)

def getConfig(section, key):
    config = configparser.ConfigParser()
    path = os.path.split(os.path.realpath(__file__))[0] + '/config.ini'
    config.read(path)
    return config.get(section, key)

def login(driver):
    driver.find_element_by_id("STAFF_ID").clear()
    driver.find_element_by_id("STAFF_ID").send_keys("账号")
    driver.find_element_by_id("LOGIN_PASSWORD").clear()
    driver.find_element_by_id("LOGIN_PASSWORD").send_keys("密码")
    driver.find_element_by_id("LOGIN_PROVINCE_CODE").find_element_by_xpath("//option[@value='97']").click()
    driver.find_element_by_id("VERIFY_CODE").clear()
    driver.implicitly_wait(10)
    driver.get_screenshot_as_file('a.png')
    location = driver.find_element_by_id('captureImage').location
    size = driver.find_element_by_id('captureImage').size
    left = location['x'] + 2
    top =  location['y'] + 2
    right = location['x'] + size['width'] + 2
    bottom = location['y'] + size['height'] + 2
    a = Image.open("a.png")
    im = a.crop((left,top,right,bottom))
    im.save("b.png")
    images = handle_split_image(im)
    v = predict(images)
    driver.find_element_by_id("VERIFY_CODE").send_keys(v)
    driver.find_element_by_xpath('//*[@id="staffLogin"]/DIV[2]/DIV/DIV/DIV[6]/INPUT[1]').click()
    # driver.implicitly_wait(10)
    # driver.find_element_by_xpath('//*[@id="staffLogin"]/DIV[2]/DIV/DIV/DIV[6]/INPUT[1]').click()
    driver.implicitly_wait(10)
    loginHtml = driver.page_source
    pattern = re.compile(r'验证码不正确')
    loginStat = pattern.findall(loginHtml)
    if len(loginStat) == 1:
        driver.implicitly_wait(20)
        driver.find_element_by_xpath('//*[@id="buttonbar"]/INPUT').click()
        return -1
    else:
        WebDriverWait(driver, 30, 3).until(
            EC.presence_of_element_located((By.ID, "notice"))
        )
        return driver

def swithJiaoFei(driver):
    driver.switch_to.default_content()
    driver.switch_to_frame('navframe')
    driver.implicitly_wait(10)
    link = driver.find_element_by_link_text(u'账务管理')
    ActionChains(driver).move_to_element(link).perform()
    driver.implicitly_wait(10)
    link = driver.find_element_by_link_text(u'账务管理')
    ActionChains(driver).move_to_element(link).perform()
    driver.find_element_by_id('FIRST_MENU_LINK_BIL6000').click()
    driver.implicitly_wait(10)
    driver.switch_to.parent_frame()
    driver.switch_to_frame('sidebarframe')
    link = driver.find_element_by_link_text(u'交费查询')
    ActionChains(driver).move_to_element(link).perform()
    driver.implicitly_wait(10)
    link = driver.find_element_by_link_text(u'交费查询')
    ActionChains(driver).move_to_element(link).perform()
    driver.find_element_by_id('BIL6512').click()
    driver.switch_to.parent_frame()
    driver.switch_to.parent_frame()
    driver.switch_to_frame('contentframe')
    driver.switch_to_frame('navframe_1')
    return driver

def Cha(driver, phone, startTime, endTime):
    """
    :param driver: 浏览器
    :param phone: 要查询的电话号码
    :param startTime: 查询开始时间
    :param endTime: 查询结束时间
    :return: 返回查询结果
    """
    driver.find_element_by_id("cond_SERIAL_NUMBER").clear()
    driver.find_element_by_id('cond_SERIAL_NUMBER').send_keys(phone)
    driver.find_element_by_id("cond_BEGIN_TIME").clear()
    driver.find_element_by_id('cond_BEGIN_TIME').send_keys(startTime)
    driver.find_element_by_id("cond_END_TIME").clear()
    driver.find_element_by_id('cond_END_TIME').send_keys(endTime)
    driver.find_element_by_id('bquerytopwithfee').click()
    # driver.switch_to_alert().accept()
    page = driver.page_source
    # soup = BeautifulSoup(page,'lxml')
    pattern = re.compile(r'查询期间内交费总额：(.*)元')
    m = pattern.search(page)
    if m == None:
        driver.implicitly_wait(10)
        driver.find_element_by_xpath('//*[@id="BUTTON_1"]').click()
        return '查询无信息'
    else:
        return m[1]

if __name__ == "__main__":
    start = datetime.datetime.now()
    url = 'https://cbss.10010.com/essframe'
    driver = webdriver.Ie()
    driver.maximize_window()
    driver.get(url)
    WebDriverWait(driver, 300).until(
        EC.presence_of_element_located((By.ID, "main"))
    )
    while True:
        loginDriver = login(driver)
        if loginDriver == -1:
            continue
        else:
            JiaoFei = swithJiaoFei(loginDriver)
            break

    # fileName = r'c:\\2.xlsx'
    fileName = getConfig("conf", "filename")
    savenum = getConfig("conf", "savenum")
    phonecol = getConfig("conf", "phonecol")
    startcol = getConfig("conf", "starttime")
    endcol = getConfig("conf", "endtime")
    resultcol = getConfig("conf", "result")
    ishead = getConfig("conf", "ishead")
    wb = Workbook()
    wb = load_workbook(fileName)
    sheetName = wb.sheetnames
    sheet = wb[sheetName[0]]
    i = 0
    if ishead == 0:
        h = 1
    else:
        h = 2
    for j in range(h, sheet.max_row + 1):
        phone = sheet.cell(row=j, column=int(phonecol)).value
        startTime = sheet.cell(row=j, column=int(startcol)).value
        if type(startTime) == str:
            startTime = startTime[:10]
        else:
            startTime = startTime.strftime("%Y-%m-%d")

        if endcol is None or endcol == '':
            endTime = datetime.datetime.now().strftime("%Y-%m-%d")
        else:
            endTime = sheet.cell(row=j, column=int(endcol)).value
            if type(endTime) == str:
                endTime = endTime[:10]
            else:
                endTime = endTime.strftime("%Y-%m-%d")

        if startTime > endTime:
            sheet["{}{}".format(resultcol, j)].value = '开始时间不能大于结束时间！'
            print('开始时间不能大于结束时间！')
        else:
            result = Cha(JiaoFei, phone, startTime, endTime )
            sheet["{}{}".format(resultcol,j)].value = result
        # print(result)
        if i >= int(savenum) :
            wb.save(fileName)
            i = 0
        else:
            i += 1
    driver.close()
    wb.save(fileName)
    end = datetime.datetime.now()

    print('查询完成，其查询数据{}条'.format(sheet.max_row))
    print('耗时 -> %s seconds' % (end - start).seconds)

## 参考址：https://www.jianshu.com/p/9bdcd62eec87   验证码
## 参考址：https://www.cnblogs.com/yoyoketang/p/6128636.html    select 下拉列表
## 参考址：https://www.jianshu.com/p/7a4414082ce2    全面简介